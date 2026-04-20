from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import subprocess

# 一、读取源文件
todo = input("请输入要处理的Excel文件（如：Cobot问题单数据.xlsx）: ")
wb1 = load_workbook(todo)
ws1 = wb1.active

# 二、获取源文档各列非空数据
seq = [cell.value for cell in ws1['A'][0:] if cell.value]
level= [cell.value for cell in ws1['D'][0:] if cell.value]
file_name = [cell.value for cell in ws1['H'][0:] if cell.value]
row_number = [cell.value for cell in ws1['I'][0:] if cell.value]
detail = [cell.value for cell in ws1['K'][0:] if cell.value]

# 创建新工作簿
wb2 = Workbook()
ws2 = wb2.active

# 写入数据
for row, (a,d,h,i,k) in enumerate(zip(seq,level,file_name,row_number,detail), start=1):
    ws2[f'A{row}'] = a
    ws2[f'C{row}'] = d
    ws2[f'D{row}'] = h
    ws2[f'E{row}'] = i
    ws2[f'G{row}'] = k
    ws2[f'F{row}'] = f'=D{row}&" 第"&E{row}&"行"'

#表头（第 1 行）
ws2['A1'] = "问题编号"
ws2['B1'] = "问题名称"
ws2['C1'] = "问题严重等级"
ws2['D1'] = "文件路径"
ws2['E1'] = "行号"
ws2['F1'] = "漏洞位置"
ws2['G1'] = "问题描述"
ws2['H1'] = "是否修改"
ws2['I1'] = "修复措施"

# 设置表头（第 1 行）全部加粗、灰色底纹
gray_fill = PatternFill(start_color='D3D3D3',end_color='D3D3D3',fill_type='solid')
for cell in ws2[1]:  # ws[1] 表示第 1 行
    cell.font = Font(bold=True)
    cell.fill = gray_fill

# 设置全文居中（遍历所有单元格）
for row in ws2.iter_rows():
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center')
# 设置所有列的宽度为6个汉字（约15个字符宽度）
for col in range(1, ws2.max_column + 1):
    ws2.column_dimensions[chr(64 + col)].width = 15
# 设置所有单元格为自动换行、水平垂直居中
for row in ws2.iter_rows():
    for cell in row:
        cell.alignment = Alignment(wrap_text=True,horizontal='center',vertical='center')
# 保存
wb2.save('Cobot过程文档.xlsx')

#打开程序
subprocess.Popen(['start', 'Cobot过程文档.xlsx'], shell=True)  # 对于Windows
# # 对于Mac: subprocess.Popen(['open', 'Cobot文档.xlsx'])
# # 对于Linux: subprocess.Popen(['xdg-open', 'Cobot文档.xlsx'])
print('成功生成文档"Cobot过程文档.xlsx"')